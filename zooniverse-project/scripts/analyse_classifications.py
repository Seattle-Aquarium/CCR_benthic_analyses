"""
analyse_classifications.py  —  Summarise a Zooniverse classification export
Version: 1.0

Produces a multi-sheet Excel report covering:
  - Overview          : file-level counts, date range, workflow list
  - Workflow Summary  : per-workflow classification & subject stats
  - Subject Summary   : per-subject retirement status, reason, vote counts
    - User Summary      : per-user classification counts, non-logged-in flag, device mix
  - Answer Breakdown  : vote distribution per answer option per workflow
  - Source Image      : classifications rolled up by source transect image
  - Time Stats        : classification duration statistics per workflow

Works on the raw Zooniverse export CSV (no prior flattening needed).

Filtering:
  --workflow-id        analyse one workflow only
  --subject-set-name   filter by subject set name stored in subject metadata
                       (Zooniverse does not embed subject_set_id in the export;
                        use source_image as a proxy with --source-image instead)
  --source-image       filter to rows whose subject source_image equals this value

Usage:
    python analyse_classifications.py

    A window opens to select:
      1. Export CSV (required)
      2. Output folder
      3. Optional workflow ID filter
      4. Optional source_image filter

Requirements:
    pip install pandas openpyxl
"""

import json
import logging
import os
import re
import subprocess
import sys
import types
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
TEAL_DARK  = "1A5276"
TEAL_MID   = "2E86C1"
TEAL_LIGHT = "D6EAF8"
TEAL_PALE  = "EBF5FB"
ORANGE     = "E67E22"
ORANGE_PAL = "FDEBD0"
GREEN_DARK = "1E8449"
GREEN_PALE = "D5F5E3"
RED_PALE   = "FADBD8"
WHITE      = "FFFFFF"
GREY_H     = "F2F3F4"


# ============================================================
# GUI INPUT FORM
# ============================================================
def get_args_via_gui():
    """Open a GUI form to collect report parameters."""
    result = {}

    root = tk.Tk()
    root.title("Analyse Zooniverse Classifications")
    root.resizable(False, False)

    pad = {"padx": 10, "pady": 5}

    ttk.Label(
        root,
        text="Classification Report Builder",
        font=("Helvetica", 13, "bold"),
    ).grid(row=0, column=0, columnspan=3, pady=(14, 4), padx=14)
    ttk.Label(
        root,
        text="Select inputs and optional filters, then click Run.",
        foreground="grey",
    ).grid(row=1, column=0, columnspan=3, pady=(0, 4))
    ttk.Separator(root, orient="horizontal").grid(
        row=2, column=0, columnspan=3, sticky="ew", padx=10, pady=4
    )

    ttk.Label(root, text="Export CSV:").grid(row=3, column=0, sticky="e", **pad)
    export_csv_var = tk.StringVar()
    ttk.Entry(root, textvariable=export_csv_var, width=55).grid(row=3, column=1, **pad)

    def browse_export_csv():
        path = filedialog.askopenfilename(
            title="Select raw Zooniverse classification export CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            export_csv_var.set(path)

    ttk.Button(root, text="Browse...", command=browse_export_csv).grid(
        row=3, column=2, **pad
    )

    ttk.Label(root, text="Output folder:").grid(row=4, column=0, sticky="e", **pad)
    output_dir_var = tk.StringVar(value="reports")
    ttk.Entry(root, textvariable=output_dir_var, width=55).grid(row=4, column=1, **pad)

    def browse_output_dir():
        path = filedialog.askdirectory(title="Select folder for Excel report")
        if path:
            output_dir_var.set(path)

    ttk.Button(root, text="Browse...", command=browse_output_dir).grid(
        row=4, column=2, **pad
    )

    ttk.Separator(root, orient="horizontal").grid(
        row=5, column=0, columnspan=3, sticky="ew", padx=10, pady=4
    )

    ttk.Label(root, text="Workflow ID (optional):").grid(row=6, column=0, sticky="e", **pad)
    workflow_id_var = tk.StringVar()
    ttk.Entry(root, textvariable=workflow_id_var, width=30).grid(row=6, column=1, sticky="w", **pad)

    ttk.Label(root, text="Source image (optional):").grid(row=7, column=0, sticky="e", **pad)
    source_image_var = tk.StringVar()
    ttk.Entry(root, textvariable=source_image_var, width=55).grid(row=7, column=1, **pad)

    open_report_var = tk.BooleanVar(value=True)
    ttk.Checkbutton(
        root,
        text="Open report when done",
        variable=open_report_var,
    ).grid(row=8, column=0, columnspan=3, sticky="w", padx=18, pady=4)

    ttk.Separator(root, orient="horizontal").grid(
        row=9, column=0, columnspan=3, sticky="ew", padx=10, pady=6
    )

    btn_frame = ttk.Frame(root)
    btn_frame.grid(row=10, column=0, columnspan=3, pady=(0, 14))

    def on_run():
        export_csv = export_csv_var.get().strip()
        output_dir = output_dir_var.get().strip()
        workflow_id_text = workflow_id_var.get().strip()
        source_image = source_image_var.get().strip()

        if not export_csv:
            messagebox.showerror("Missing input", "Please select an export CSV file.")
            return
        if not Path(export_csv).is_file():
            messagebox.showerror("File not found", f"Export CSV not found:\n{export_csv}")
            return
        if not output_dir:
            messagebox.showerror("Missing input", "Please select an output folder.")
            return

        if workflow_id_text:
            try:
                workflow_id = int(workflow_id_text)
            except ValueError:
                messagebox.showerror("Invalid input", "Workflow ID must be a whole number.")
                return
        else:
            workflow_id = None

        result["export_csv"] = export_csv
        result["output_dir"] = output_dir
        result["workflow_id"] = workflow_id
        result["source_image"] = source_image or None
        result["open_report"] = open_report_var.get()
        result["submitted"] = True
        root.destroy()

    def on_cancel():
        root.destroy()

    ttk.Button(btn_frame, text="  Run  ", command=on_run).pack(side="left", padx=8)
    ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side="left", padx=8)

    root.mainloop()

    if not result.get("submitted"):
        print("Cancelled by user.")
        sys.exit(0)

    return types.SimpleNamespace(
        export_csv=result["export_csv"],
        output_dir=result["output_dir"],
        workflow_id=result["workflow_id"],
        source_image=result["source_image"],
        open_report=result["open_report"],
    )


# ============================================================
# PARSING HELPERS
# ============================================================
IMG_TAG_RE = re.compile(r'!\[.*?\]\(.*?\)\s*')

def clean_answer(val: str) -> str:
    """Strip Zooniverse markdown image tags from answer labels."""
    if not isinstance(val, str):
        return str(val) if val is not None else ""
    return IMG_TAG_RE.sub("", val).strip()


def parse_json_col(series: pd.Series, col_name: str) -> list[dict | None]:
    """Parse a column of JSON strings; return list of dicts (None on failure)."""
    out = []
    for val in series:
        try:
            out.append(json.loads(val))
        except Exception:
            out.append(None)
    return out


def parse_metadata(series: pd.Series) -> pd.DataFrame:
    """Extract started_at, finished_at, device type, user_group membership."""
    rows = []
    for val in series:
        rec = {"duration_s": None, "device": "unknown", "in_user_group": False}
        try:
            m = json.loads(val)
            s = pd.Timestamp(m["started_at"])
            f = pd.Timestamp(m["finished_at"])
            rec["duration_s"] = max(0, (f - s).total_seconds())
            ua = m.get("user_agent", "")
            rec["device"] = "mobile" if any(
                t in ua for t in ("Mobile", "Android", "iPhone", "iPad")
            ) else "desktop"
            rec["in_user_group"] = len(m.get("user_group_ids", [])) > 0
        except Exception:
            pass
        rows.append(rec)
    return pd.DataFrame(rows)


def parse_subject_data(series: pd.Series) -> pd.DataFrame:
    """
    Flatten subject_data JSON into one row per classification.
    Extracts: subject_id, retired (bool), retirement_reason,
              classifications_count_at_retirement, source_image,
              filename, model_pred_code, model_pred_name.
    """
    rows = []
    for val in series:
        rec = {
            "subject_id_extracted":             None,
            "is_retired":                       False,
            "retirement_reason":                None,
            "classifications_count_at_retire":  None,
            "retired_at":                       None,
            "source_image":                     None,
            "filename":                         None,
            "model_pred_code":                  None,
            "model_pred_name":                  None,
        }
        try:
            sd = json.loads(val)
            for subj_id, meta in sd.items():
                rec["subject_id_extracted"] = subj_id
                r = meta.get("retired")
                if r and isinstance(r, dict):
                    rec["is_retired"]                      = True
                    rec["retirement_reason"]               = r.get("retirement_reason")
                    rec["classifications_count_at_retire"] = r.get("classifications_count")
                    rec["retired_at"]                      = r.get("retired_at")
                rec["source_image"]     = meta.get("source_image")
                rec["filename"]         = meta.get("filename")
                rec["model_pred_code"]  = meta.get("model_pred_code")
                rec["model_pred_name"]  = meta.get("model_pred_name")
        except Exception:
            pass
        rows.append(rec)
    return pd.DataFrame(rows)


def parse_annotations(series: pd.Series) -> pd.Series:
    """Return the first task answer for each classification (cleaned)."""
    answers = []
    for val in series:
        try:
            ann = json.loads(val)
            if ann and isinstance(ann, list):
                answers.append(clean_answer(ann[0].get("value", "")))
            else:
                answers.append("")
        except Exception:
            answers.append("")
    return pd.Series(answers, dtype=str)


# ============================================================
# LOAD & FLATTEN
# ============================================================
def load_and_flatten(csv_path: str,
                     workflow_id: int | None,
                     source_image_filter: str | None) -> pd.DataFrame:
    """
    Load the raw export CSV and join all parsed columns into a single flat table.
    Applies --workflow-id and --source-image filters if provided.
    """
    log.info(f"Loading {csv_path}…")
    raw = pd.read_csv(csv_path, low_memory=False)
    log.info(f"Loaded {len(raw):,} rows, {len(raw.columns)} columns")

    # Workflow filter
    if workflow_id is not None:
        raw = raw[raw["workflow_id"] == workflow_id].copy()
        log.info(f"After workflow filter ({workflow_id}): {len(raw):,} rows")
        if raw.empty:
            log.error(f"No rows found for workflow_id={workflow_id}. "
                      f"Available: {list(raw['workflow_id'].unique())}")
            sys.exit(1)

    log.info("Parsing JSON columns…")
    subj_df = parse_subject_data(raw["subject_data"])
    meta_df = parse_metadata(raw["metadata"])
    answers  = parse_annotations(raw["annotations"])

    flat = pd.concat([
        raw[["classification_id", "user_name", "user_id",
             "workflow_id", "workflow_name", "workflow_version",
             "created_at", "subject_ids"]].reset_index(drop=True),
        subj_df.reset_index(drop=True),
        meta_df.reset_index(drop=True),
        answers.rename("answer").reset_index(drop=True),
    ], axis=1)

    # Derived columns
    user_name_norm = flat["user_name"].fillna("").astype(str).str.strip().str.lower()
    flat["is_anonymous"]    = user_name_norm.str.startswith("not-logged-in")
    flat["created_at"]      = pd.to_datetime(flat["created_at"], utc=True, errors="coerce")

    # Source-image filter (applied after parsing)
    if source_image_filter is not None:
        flat = flat[flat["source_image"] == source_image_filter].copy()
        log.info(f"After source_image filter ('{source_image_filter}'): {len(flat):,} rows")
        if flat.empty:
            log.error(f"No rows matched source_image='{source_image_filter}'.")
            sys.exit(1)

    log.info(f"Flat table: {len(flat):,} rows, {len(flat.columns)} columns")
    return flat


# ============================================================
# SUMMARY BUILDERS
# ============================================================
def build_overview(flat: pd.DataFrame, filters: dict) -> pd.DataFrame:
    wf_list = (flat[["workflow_id", "workflow_name"]]
               .drop_duplicates()
               .assign(combined=lambda d: d["workflow_id"].astype(str)
                       + " — " + d["workflow_name"])
               ["combined"].tolist())

    rows = [
        ("Export file rows analysed",   len(flat)),
        ("Unique subjects",             flat["subject_ids"].nunique()),
        ("Unique classifiers (named)",  flat.loc[~flat["is_anonymous"], "user_name"].nunique()),
        ("Non-logged-in classifications", flat["is_anonymous"].sum()),
        ("Non-logged-in %",             f"{flat['is_anonymous'].mean()*100:.1f}%"),
        ("Retired subjects",            flat.drop_duplicates("subject_ids")["is_retired"].sum()),
        ("Not yet retired subjects",    (~flat.drop_duplicates("subject_ids")["is_retired"]).sum()),
        ("Date range (first)",          str(flat["created_at"].min())[:19] if flat["created_at"].notna().any() else "—"),
        ("Date range (last)",           str(flat["created_at"].max())[:19] if flat["created_at"].notna().any() else "—"),
        ("Workflows in this file",      ", ".join(wf_list)),
        ("— filter: workflow_id",       str(filters.get("workflow_id") or "all")),
        ("— filter: source_image",      str(filters.get("source_image") or "all")),
        ("Report generated",            datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def build_workflow_summary(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for wf_id, g in flat.groupby("workflow_id"):
        subjects = g.drop_duplicates("subject_ids")
        retired  = subjects[subjects["is_retired"]]
        rows.append({
            "Workflow ID":                  wf_id,
            "Workflow Name":                g["workflow_name"].iloc[0],
            "Workflow Version(s)":          ", ".join(g["workflow_version"].astype(str).unique()),
            "Total Classifications":        len(g),
            "Unique Subjects":              g["subject_ids"].nunique(),
            "Retired Subjects":             len(retired),
            "Not Retired":                  len(subjects) - len(retired),
            "Retired — consensus":          (retired["retirement_reason"] == "consensus").sum(),
            "Retired — classification_count": (retired["retirement_reason"] == "classification_count").sum(),
            "Retired — other":              (~retired["retirement_reason"].isin(
                                                ["consensus", "classification_count"])).sum()
                                            if len(retired) else 0,
            "Avg classifications/subject":  f"{g.groupby('subject_ids').size().mean():.2f}",
            "Non-logged-in classifications": g["is_anonymous"].sum(),
            "Non-logged-in %":            f"{g['is_anonymous'].mean()*100:.1f}%",
            "Unique named users":           g.loc[~g["is_anonymous"], "user_name"].nunique(),
            "Desktop classifications":      (g["device"] == "desktop").sum(),
            "Mobile classifications":       (g["device"] == "mobile").sum(),
            "Users in a group":             g["in_user_group"].sum(),
        })
    return pd.DataFrame(rows)


def build_subject_summary(flat: pd.DataFrame) -> pd.DataFrame:
    """One row per unique subject with retirement info and vote counts."""
    rows = []
    for subj_id, g in flat.groupby("subject_ids"):
        # Retirement info comes from any row for this subject (all carry same data)
        r      = g.iloc[0]
        retired = r["is_retired"]

        # Vote tally per answer
        vote_counts = g["answer"].value_counts().to_dict()
        top_answer  = g["answer"].mode().iloc[0] if not g["answer"].empty else ""
        n_answers   = g["answer"].nunique()

        rows.append({
            "Subject ID":                       subj_id,
            "Filename":                         r["filename"],
            "Source Image":                     r["source_image"],
            "Model Prediction Code":            r["model_pred_code"],
            "Model Prediction Name":            r["model_pred_name"],
            "Total Classifications":            len(g),
            "Unique Answer Options Seen":       n_answers,
            "Top Answer (most votes)":          top_answer,
            "Is Retired":                       retired,
            "Retirement Reason":                r["retirement_reason"] if retired else "—",
            "Classifications at Retirement":    r["classifications_count_at_retire"] if retired else "—",
            "Retired At":                       str(r["retired_at"])[:19] if retired and r["retired_at"] else "—",
            "Non-logged-in Classifications":    g["is_anonymous"].sum(),
            "Named User Classifications":       (~g["is_anonymous"]).sum(),
            "Workflow(s)":                      ", ".join(g["workflow_id"].astype(str).unique()),
            **{f"Votes — {k}": v for k, v in vote_counts.items()},
        })
    df = pd.DataFrame(rows)
    # Sort: retired first, then by total classifications desc
    df = df.sort_values(["Is Retired", "Total Classifications"],
                        ascending=[False, False]).reset_index(drop=True)
    return df


def build_user_summary(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for user, g in flat.groupby("user_name"):
        rows.append({
            "User Name":                user,
            "Is Non-logged-in":         g["is_anonymous"].iloc[0],
            "Total Classifications":    len(g),
            "Workflows Contributed To": ", ".join(g["workflow_id"].astype(str).unique()),
            "Unique Subjects Seen":     g["subject_ids"].nunique(),
            "In a User Group":          g["in_user_group"].any(),
            "Desktop":                  (g["device"] == "desktop").sum(),
            "Mobile":                   (g["device"] == "mobile").sum(),
            "Avg Duration (s)":         f"{g['duration_s'].mean():.1f}" if g["duration_s"].notna().any() else "—",
            "Min Duration (s)":         f"{g['duration_s'].min():.1f}" if g["duration_s"].notna().any() else "—",
            "First Classification":     str(g["created_at"].min())[:19],
            "Last Classification":      str(g["created_at"].max())[:19],
        })
    df = pd.DataFrame(rows).sort_values("Total Classifications", ascending=False).reset_index(drop=True)
    return df


def build_answer_breakdown(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for wf_id, g in flat.groupby("workflow_id"):
        wf_name   = g["workflow_name"].iloc[0]
        total     = len(g)
        vc        = g["answer"].value_counts()
        for answer, count in vc.items():
            rows.append({
                "Workflow ID":          wf_id,
                "Workflow Name":        wf_name,
                "Answer":               answer,
                "Count":                count,
                "% of Workflow Total":  f"{count/total*100:.1f}%",
            })
    return pd.DataFrame(rows)


def build_source_image_summary(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for src, g in flat.groupby("source_image"):
        subjects  = g.drop_duplicates("subject_ids")
        retired   = subjects[subjects["is_retired"]]
        rows.append({
            "Source Image":                 src,
            "Total Subjects":               len(subjects),
            "Retired":                      len(retired),
            "Not Retired":                  len(subjects) - len(retired),
            "Retired — consensus":          (retired["retirement_reason"] == "consensus").sum(),
            "Retired — classification_count": (retired["retirement_reason"] == "classification_count").sum(),
            "Total Classifications":        len(g),
            "Avg Classifications/Subject":  f"{len(g)/len(subjects):.2f}" if len(subjects) else "—",
            "Unique Users":                 g.loc[~g["is_anonymous"], "user_name"].nunique(),
            "Non-logged-in Classifications": g["is_anonymous"].sum(),
            "Workflow(s)":                  ", ".join(g["workflow_id"].astype(str).unique()),
        })
    df = pd.DataFrame(rows).sort_values("Total Subjects", ascending=False).reset_index(drop=True)
    return df


def build_time_stats(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for wf_id, g in flat.groupby("workflow_id"):
        d = g["duration_s"].dropna()
        if d.empty:
            continue
        rows.append({
            "Workflow ID":          wf_id,
            "Workflow Name":        g["workflow_name"].iloc[0],
            "N (with timing)":      len(d),
            "Mean (s)":             round(d.mean(), 1),
            "Median (s)":           round(d.median(), 1),
            "Std Dev (s)":          round(d.std(), 1),
            "Min (s)":              round(d.min(), 1),
            "Max (s)":              round(d.max(), 1),
            "< 3 s (likely rushed)": (d < 3).sum(),
            "< 3 s %":              f"{(d < 3).mean()*100:.1f}%",
            "> 60 s":               (d > 60).sum(),
        })
    return pd.DataFrame(rows)


# ============================================================
# EXCEL FORMATTING
# ============================================================
def _thin():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    thin = Side(style="thin",   color="BDBDBD")
    med  = Side(style="medium", color=TEAL_MID)
    return Border(left=thin, right=thin, top=thin, bottom=med)

def _fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def _font(bold=False, size=10, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def style_sheet(ws, df: pd.DataFrame,
                header_color: str = TEAL_MID,
                title: str = "",
                col_widths: dict | None = None):
    """Write df to ws with consistent header + alternating row styling."""
    ws.sheet_view.showGridLines = False

    start_row = 1
    if title:
        ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
        c = ws["A1"]
        c.value     = title
        c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        c.fill      = _fill(TEAL_DARK)
        c.alignment = _center()
        ws.row_dimensions[1].height = 26
        start_row = 2

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=col_idx, value=col_name)
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill(header_color)
        c.alignment = _center(wrap=True)
        c.border    = _thick_bottom()
    ws.row_dimensions[start_row].height = 28

    # Data rows
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        bg = TEAL_PALE if r_idx % 2 == 0 else WHITE
        for c_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_idx, column=c_idx,
                        value=None if (isinstance(val, float) and pd.isna(val)) else val)
            c.font      = _font()
            c.fill      = _fill(bg)
            c.alignment = _left(wrap=False)
            c.border    = _thin()
        ws.row_dimensions[r_idx].height = 16

    # Column widths
    default_widths = {i: max(len(str(col)) + 4, 12)
                      for i, col in enumerate(df.columns, start=1)}
    # Widen based on data sample
    for r_idx, row_data in enumerate(df.head(40).itertuples(index=False)):
        for c_idx, val in enumerate(row_data, start=1):
            w = min(len(str(val)) + 3, 50)
            if w > default_widths.get(c_idx, 0):
                default_widths[c_idx] = w

    for c_idx, width in default_widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    if col_widths:
        for col_name, width in col_widths.items():
            if col_name in df.columns:
                c_idx = list(df.columns).index(col_name) + 1
                ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ============================================================
# WRITE REPORT
# ============================================================
def write_report(sheets: dict[str, pd.DataFrame], output_path: Path,
                 header_colors: dict[str, str] | None = None):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    colors = header_colors or {}
    titles = {
        "Overview":         "📊  Classification Export — Overview",
        "Workflow Summary":  "📋  Workflow Summary",
        "Subject Summary":   "🔎  Subject Summary",
        "User Summary":      "👤  User Summary",
        "Answer Breakdown":  "✅  Answer Breakdown",
        "Source Image":      "🗂️  Source Image Summary",
        "Time Stats":        "⏱️  Classification Time Statistics",
    }

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name)
        style_sheet(
            ws, df,
            header_color=colors.get(sheet_name, TEAL_MID),
            title=titles.get(sheet_name, sheet_name),
        )
        log.info(f"  Sheet '{sheet_name}': {len(df)} rows")

    wb.save(output_path)
    log.info(f"Report saved → {output_path}")


# ============================================================
# ENTRY POINT
# ============================================================
def main():
    args = get_args_via_gui()

    csv_path = Path(args.export_csv)
    if not csv_path.is_file():
        log.error(f"Export CSV not found: {csv_path}")
        sys.exit(1)

    filters = {
        "workflow_id":  args.workflow_id,
        "source_image": args.source_image,
    }

    flat = load_and_flatten(
        str(csv_path),
        workflow_id=args.workflow_id,
        source_image_filter=args.source_image,
    )

    log.info("Building summary sheets…")
    sheets = {
        "Overview":         build_overview(flat, filters),
        "Workflow Summary":  build_workflow_summary(flat),
        "Subject Summary":   build_subject_summary(flat),
        "User Summary":      build_user_summary(flat),
        "Answer Breakdown":  build_answer_breakdown(flat),
        "Source Image":      build_source_image_summary(flat),
        "Time Stats":        build_time_stats(flat),
    }

    sheet_colors = {
        "Overview":         TEAL_DARK,
        "Workflow Summary":  TEAL_MID,
        "Subject Summary":   TEAL_MID,
        "User Summary":      "5D6D7E",
        "Answer Breakdown":  GREEN_DARK,
        "Source Image":      ORANGE,
        "Time Stats":        "7D3C98",
    }

    # Build output filename
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str   = datetime.now().strftime("%Y%m%d_%H%M")
    wf_suffix  = f"_wf{args.workflow_id}" if args.workflow_id else ""
    si_suffix  = f"_{args.source_image}"  if args.source_image else ""
    out_path   = output_dir / f"classification_report{wf_suffix}{si_suffix}_{date_str}.xlsx"

    write_report(sheets, out_path, header_colors=sheet_colors)

    if args.open_report:
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(out_path))
            elif sys.platform == "darwin":
                subprocess.run(["open", str(out_path)], check=False)
            else:
                subprocess.run(["xdg-open", str(out_path)], check=False)
        except Exception as e:
            log.warning(f"Could not open report automatically: {e}")

    # Console summary
    log.info("=" * 60)
    log.info("✅  Report complete!")
    log.info(f"    Rows analysed     : {len(flat):,}")
    log.info(f"    Unique subjects   : {flat['subject_ids'].nunique():,}")
    log.info(f"    Retired subjects  : {flat.drop_duplicates('subject_ids')['is_retired'].sum():,}")
    log.info(f"    Unique workflows  : {flat['workflow_id'].nunique()}")
    log.info(f"    Output            : {out_path}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
