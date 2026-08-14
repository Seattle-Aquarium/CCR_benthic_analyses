"""
analyse_classifications.py  —  Summarise a Zooniverse classification export
Version: 1.2

Produces a multi-sheet Excel report covering:
  - Overview            : file-level counts, date range, workflow list, confirmation-rule summary
  - Workflow Summary    : per-workflow classification & subject stats
  - Subject Summary     : per-subject retirement status, reason, vote counts, Outcome
  - User Summary        : per-user classification counts, non-logged-in flag, device mix
  - Answer Breakdown    : vote distribution per answer option per workflow
  - Source Image        : classifications rolled up by source transect image, with
                           per-image patch counts for Verified / Needs Toolbox Review /
                           Zooniverse — Needs More Votes
  - Time Stats          : classification duration statistics per workflow
  - Transect Completion : per-transect (source_image) progress through the two-workflow pipeline
      Yes/No workflows (30787, 32022, 31534): subjects confirmed (yes) or denied (no/not sure)
      Multi workflows  (30752, 32023, 31535): subjects denied in yes/no are classified here

Outcome status — three buckets, and "Toolbox" is reserved for exactly one:
  - Verified …            : Zooniverse crowd/expert consensus made the call. This
                             is what gets written into Toolbox's Verified=TRUE field,
                             but the decision itself happened on Zooniverse.
  - Zooniverse — Needs More Votes : still active on Zooniverse, waiting on the
                             crowd/expert queue — not a Toolbox concern, may
                             resolve on its own.
  - Needs Toolbox Review   : Zooniverse is done and produced no verdict (explicit
                             "not sure", or retired from a pipeline stage without
                             reaching threshold) — bring it into Toolbox for a
                             human/expert to decide.
  "Outcome" / "consensus_status" (Subject Summary, Source Image, Transect
  Completion) is computed with the exact same vote-threshold rules and priority
  order used by zooni_to_toolbox_annot.py to decide Verified vs Label="Review" —
  see compute_consensus(). This is intentionally distinct from "Is Retired
  (Zooniverse)", which only reflects the platform's own retirement rules
  (e.g. classification_count) and can disagree with the confirmation outcome.

Works on the raw Zooniverse export CSV (no prior flattening needed).

Filtering:
  --workflow-id        analyse one workflow only
  --subject-set-name   filter by subject set name stored in subject metadata
                       (Zooniverse does not embed subject_set_id in the export;
                        use source_image as a proxy with --source-image instead)
  --source-image       filter to rows whose subject source_image equals this value

Usage:
    python analyse_classifications.py

    A window opens to select:
      1. Export CSV (required)
      2. Output folder
      3. Optional workflow ID filter
      4. Optional source_image filter

Requirements:
    pip install pandas openpyxl
"""

import json
import logging
import os
import re
import subprocess
import sys
import types
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
TEAL_DARK  = "1A5276"
TEAL_MID   = "2E86C1"
TEAL_LIGHT = "D6EAF8"
TEAL_PALE  = "EBF5FB"
ORANGE     = "E67E22"
ORANGE_PAL = "FDEBD0"
GREEN_DARK = "1E8449"
GREEN_PALE = "D5F5E3"
RED_PALE   = "FADBD8"
WHITE      = "FFFFFF"
GREY_H     = "F2F3F4"

# ── Workflow IDs ──────────────────────────────────────────────────────────────
YESNO_WORKFLOW_IDS        = {30787, 32022}   # confirm/deny workflows
YESNO_EXPERT_WORKFLOW_IDS = {31534}          # expert confirm/deny
MULTI_WORKFLOW_IDS        = {30752, 32023}   # multiple-choice classification workflows
MULTI_EXPERT_WORKFLOW_IDS = {31535}          # expert multiple-choice

# ── Confirmation thresholds ─────────────────────────────────────────────────
# Must mirror zooni_to_toolbox_annot.py exactly — "Verified" here is a
# Zooniverse-side consensus decision and means the same thing "Verified"
# means once written into the Toolbox import; "Needs Toolbox Review" here
# means the row will land with Label="Review" in Toolbox.
YESNO_AGREE_MIN_N    = 5
YESNO_AGREE_MIN_FRAC = 0.75
MULTI_AGREE_MIN_N    = 3
MULTI_AGREE_MIN_FRAC = 0.67
EXPERT_MIN_N         = 1

NOT_SURE_RE = re.compile(r"not sure", re.IGNORECASE)

# consensus_status -> (short outcome label, detail text).
#
# Three buckets, and the "Toolbox" word is reserved for exactly one of them:
#   - Verified …           : Zooniverse crowd/expert consensus made the call.
#                             Nothing for a person to do — this is what gets
#                             written into Toolbox's Verified=TRUE field, but
#                             the decision itself happened on Zooniverse, not
#                             in Toolbox.
#   - Zooniverse — Needs More Votes : still active on Zooniverse, waiting on
#                             the crowd/expert queue. Also not a Toolbox
#                             concern — it may still resolve on its own.
#   - Needs Toolbox Review  : Zooniverse is done and produced no verdict
#                             (explicit "not sure", or retired from a stage
#                             without reaching threshold). This is the only
#                             bucket that means "bring it into Toolbox for a
#                             human/expert to decide."
#
# Being denied in Yes/No is NOT by itself a reason for Toolbox review — a
# "No" consensus is exactly what routes a subject into the Multi-choice
# workflow next. It only becomes "Needs Toolbox Review" once that next stage
# has also run its course (retired without reaching Multi-choice consensus)
# or has returned an explicit "not sure". Until then it's
# "pending_multi_review": still on Zooniverse's pipeline, no human action
# needed yet.
CONSENSUS_OUTCOME_MAP = {
    "multi_expert":     ("Verified — Multi-choice (expert)", "Expert multi-choice consensus"),
    "multi_consensus":  ("Verified — Multi-choice (crowd)",  "Volunteer multi-choice consensus"),
    "confirm_expert":   ("Verified — Yes/No (expert)",       "Confirmed by expert yes/no vote"),
    "confirm_pred":     ("Verified — Yes/No (crowd)",        "Confirmed by volunteer yes/no vote"),
    "voted_review":     ("Needs Toolbox Review",             "Volunteers voted 'not sure' in Multi-choice — "
                                                               "bring into Toolbox for expert review"),
    "stalled_no_consensus": ("Needs Toolbox Review",         "Retired on Zooniverse without reaching the "
                                                               "confirmation threshold — no more votes are coming; "
                                                               "bring into Toolbox for expert review"),
    "pending_multi_review": ("Zooniverse — Needs More Votes", "Denied in Yes/No; awaiting Multi-choice consensus "
                                                               "on Zooniverse"),
    "needs_more_votes": ("Zooniverse — Needs More Votes",    "Still active on Zooniverse and may reach consensus "
                                                               "on its own"),
}


# ============================================================
# GUI INPUT FORM
# ============================================================
def get_args_via_gui():
    """Open a GUI form to collect report parameters."""
    result = {}

    root = tk.Tk()
    root.title("Analyse Zooniverse Classifications")
    root.resizable(False, False)

    pad = {"padx": 10, "pady": 5}

    ttk.Label(
        root,
        text="Classification Report Builder",
        font=("Helvetica", 13, "bold"),
    ).grid(row=0, column=0, columnspan=3, pady=(14, 4), padx=14)
    ttk.Label(
        root,
        text="Select inputs and optional filters, then click Run.",
        foreground="grey",
    ).grid(row=1, column=0, columnspan=3, pady=(0, 4))
    ttk.Separator(root, orient="horizontal").grid(
        row=2, column=0, columnspan=3, sticky="ew", padx=10, pady=4
    )

    ttk.Label(root, text="Export CSV:").grid(row=3, column=0, sticky="e", **pad)
    export_csv_var = tk.StringVar()
    ttk.Entry(root, textvariable=export_csv_var, width=55).grid(row=3, column=1, **pad)

    def browse_export_csv():
        path = filedialog.askopenfilename(
            title="Select raw Zooniverse classification export CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if path:
            export_csv_var.set(path)

    ttk.Button(root, text="Browse...", command=browse_export_csv).grid(
        row=3, column=2, **pad
    )

    ttk.Label(root, text="Output folder:").grid(row=4, column=0, sticky="e", **pad)
    output_dir_var = tk.StringVar()
    ttk.Entry(root, textvariable=output_dir_var, width=55).grid(row=4, column=1, **pad)

    def browse_output_dir():
        path = filedialog.askdirectory(title="Select folder for Excel report")
        if path:
            output_dir_var.set(path)

    ttk.Button(root, text="Browse...", command=browse_output_dir).grid(
        row=4, column=2, **pad
    )

    ttk.Separator(root, orient="horizontal").grid(
        row=5, column=0, columnspan=3, sticky="ew", padx=10, pady=4
    )

    ttk.Label(root, text="Workflow ID (optional):").grid(row=6, column=0, sticky="e", **pad)
    workflow_id_var = tk.StringVar()
    ttk.Entry(root, textvariable=workflow_id_var, width=30).grid(row=6, column=1, sticky="w", **pad)

    ttk.Label(root, text="Source image (optional):").grid(row=7, column=0, sticky="e", **pad)
    source_image_var = tk.StringVar()
    ttk.Entry(root, textvariable=source_image_var, width=55).grid(row=7, column=1, **pad)

    ttk.Label(root, text="Transect ID (optional):").grid(row=8, column=0, sticky="e", **pad)
    transect_id_var = tk.StringVar()
    ttk.Entry(root, textvariable=transect_id_var, width=55).grid(row=8, column=1, **pad)
    ttk.Label(root, text="Comma-separated for multiple, e.g. T01, T02, T03",
              foreground="grey").grid(row=8, column=2, sticky="w")

    # Date range filter
    date_frame = ttk.Frame(root)
    date_frame.grid(row=9, column=0, columnspan=3, sticky="w", padx=10, pady=2)
    ttk.Label(date_frame, text="Classification date (optional):").grid(
        row=0, column=0, sticky="e", padx=(0, 6))
    date_from_var = tk.StringVar()
    ttk.Entry(date_frame, textvariable=date_from_var, width=14).grid(row=0, column=1, sticky="w")
    ttk.Label(date_frame, text="to").grid(row=0, column=2, padx=6)
    date_to_var = tk.StringVar()
    ttk.Entry(date_frame, textvariable=date_to_var, width=14).grid(row=0, column=3, sticky="w")
    ttk.Label(date_frame, text="(YYYY-MM-DD)", foreground="grey").grid(
        row=0, column=4, sticky="w", padx=(8, 0))

    open_report_var = tk.BooleanVar(value=True)
    ttk.Checkbutton(
        root,
        text="Open report when done",
        variable=open_report_var,
    ).grid(row=10, column=0, columnspan=3, sticky="w", padx=18, pady=4)

    ttk.Separator(root, orient="horizontal").grid(
        row=11, column=0, columnspan=3, sticky="ew", padx=10, pady=6
    )

    btn_frame = ttk.Frame(root)
    btn_frame.grid(row=12, column=0, columnspan=3, pady=(0, 14))

    def on_run():
        export_csv = export_csv_var.get().strip()
        output_dir = output_dir_var.get().strip()
        workflow_id_text = workflow_id_var.get().strip()
        source_image = source_image_var.get().strip()
        transect_id  = transect_id_var.get().strip()
        date_from    = date_from_var.get().strip()
        date_to      = date_to_var.get().strip()

        if not export_csv:
            messagebox.showerror("Missing input", "Please select an export CSV file.")
            return
        if not Path(export_csv).is_file():
            messagebox.showerror("File not found", f"Export CSV not found:\n{export_csv}")
            return
        if not output_dir:
            messagebox.showerror("Missing input", "Please select an output folder.")
            return

        if workflow_id_text:
            try:
                workflow_id = int(workflow_id_text)
            except ValueError:
                messagebox.showerror("Invalid input", "Workflow ID must be a whole number.")
                return
        else:
            workflow_id = None

        result["export_csv"] = export_csv
        result["output_dir"] = output_dir
        result["workflow_id"] = workflow_id
        result["source_image"] = source_image or None
        result["transect_id"]  = transect_id or None
        result["date_from"]    = date_from or None
        result["date_to"]      = date_to or None
        result["open_report"]  = open_report_var.get()
        result["submitted"] = True
        root.destroy()

    def on_cancel():
        root.destroy()

    ttk.Button(btn_frame, text="  Run  ", command=on_run).pack(side="left", padx=8)
    ttk.Button(btn_frame, text="Cancel", command=on_cancel).pack(side="left", padx=8)

    root.mainloop()

    if not result.get("submitted"):
        print("Cancelled by user.")
        sys.exit(0)

    return types.SimpleNamespace(
        export_csv=result["export_csv"],
        output_dir=result["output_dir"],
        workflow_id=result["workflow_id"],
        source_image=result["source_image"],
        transect_id=result["transect_id"],
        date_from=result["date_from"],
        date_to=result["date_to"],
        open_report=result["open_report"],
    )


# ============================================================
# PARSING HELPERS
# ============================================================
IMG_TAG_RE = re.compile(r'!\[.*?\]\(.*?\)\s*')

def clean_answer(val: str) -> str:
    """Strip Zooniverse markdown image tags from answer labels."""
    if not isinstance(val, str):
        return str(val) if val is not None else ""
    return IMG_TAG_RE.sub("", val).strip()


def parse_json_col(series: pd.Series, col_name: str) -> list[dict | None]:
    """Parse a column of JSON strings; return list of dicts (None on failure)."""
    out = []
    for val in series:
        try:
            out.append(json.loads(val))
        except Exception:
            out.append(None)
    return out


def parse_metadata(series: pd.Series) -> pd.DataFrame:
    """Extract started_at, finished_at, device type, user_group membership."""
    rows = []
    for val in series:
        rec = {"duration_s": None, "device": "unknown", "in_user_group": False}
        try:
            m = json.loads(val)
            s = pd.Timestamp(m["started_at"])
            f = pd.Timestamp(m["finished_at"])
            rec["duration_s"] = max(0, (f - s).total_seconds())
            ua = m.get("user_agent", "")
            rec["device"] = "mobile" if any(
                t in ua for t in ("Mobile", "Android", "iPhone", "iPad")
            ) else "desktop"
            rec["in_user_group"] = len(m.get("user_group_ids", [])) > 0
        except Exception:
            pass
        rows.append(rec)
    return pd.DataFrame(rows)


def parse_subject_data(series: pd.Series) -> pd.DataFrame:
    """
    Flatten subject_data JSON into one row per classification.
    Extracts: subject_id, retired (bool), retirement_reason,
              classifications_count_at_retirement, source_image,
              transect_id, filename, model_pred_code, model_pred_name.
    """
    rows = []
    for val in series:
        rec = {
            "subject_id_extracted":             None,
            "is_retired":                       False,
            "retirement_reason":                None,
            "classifications_count_at_retire":  None,
            "retired_at":                       None,
            "source_image":                     None,
            "transect_id":                      None,
            "filename":                         None,
            "model_pred_code":                  None,
            "model_pred_name":                  None,
        }
        try:
            sd = json.loads(val)
            for subj_id, meta in sd.items():
                rec["subject_id_extracted"] = subj_id
                r = meta.get("retired")
                if r and isinstance(r, dict):
                    rec["is_retired"]                      = True
                    rec["retirement_reason"]               = r.get("retirement_reason")
                    rec["classifications_count_at_retire"] = r.get("classifications_count")
                    rec["retired_at"]                      = r.get("retired_at")
                rec["source_image"]     = meta.get("source_image")
                rec["transect_id"]      = meta.get("transect_id")
                rec["filename"]         = meta.get("filename")
                rec["model_pred_code"]  = meta.get("model_pred_code")
                rec["model_pred_name"]  = meta.get("model_pred_name")
        except Exception:
            pass
        rows.append(rec)
    return pd.DataFrame(rows)


def _annotation_value_to_string(v) -> str:
    """Coerce a Zooniverse annotation `value` (str/num/bool/list/dict) to text."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float, bool)):
        return str(v)
    if isinstance(v, list):
        return ", ".join(str(x).strip() for x in v if str(x).strip())
    if isinstance(v, dict):
        for key in ("choice", "label", "value", "answers"):
            if key in v:
                s = _annotation_value_to_string(v.get(key))
                if s:
                    return s
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    return str(v).strip()


def parse_annotations(series: pd.Series, workflow_ids: pd.Series) -> pd.Series:
    """
    Return the answer used for consensus for each classification.

    Yes/No workflows: the single task's answer (cleaned).
    Multi-choice workflows: most-specific answered task wins (T3 > T2 > T1 > T0),
    mirroring extract_multi_chosen_label() in zooni_to_toolbox_annot.py so that
    "top answer" here means the same thing as the consensus label that drives
    the Toolbox import — a follow-up sub-choice (T1/T2/T3), not the broad T0
    category, decides the vote.
    """
    multi_ids = MULTI_WORKFLOW_IDS | MULTI_EXPERT_WORKFLOW_IDS
    answers = []
    for val, wf_id in zip(series, workflow_ids):
        try:
            ann = json.loads(val)
        except Exception:
            answers.append("")
            continue
        if not isinstance(ann, list) or not ann:
            answers.append("")
            continue

        if wf_id in multi_ids:
            by_task = {d.get("task"): d.get("value", "")
                       for d in ann if isinstance(d, dict) and "task" in d}
            chosen = ""
            for t in ("T3", "T2", "T1", "T0"):
                v = _annotation_value_to_string(by_task.get(t, ""))
                if v:
                    chosen = v
                    break
            answers.append(clean_answer(chosen))
        else:
            answers.append(clean_answer(_annotation_value_to_string(ann[0].get("value", ""))))
    return pd.Series(answers, dtype=str)


# ============================================================
# CONSENSUS (mirrors zooni_to_toolbox_annot.py)
# ============================================================
def compute_consensus(flat: pd.DataFrame) -> pd.DataFrame:
    """
    Per-subject confirmation status, using the exact workflow groupings and
    vote thresholds from zooni_to_toolbox_annot.py, extended to reflect the
    two-stage Zooniverse pipeline (Yes/No, then Multi-choice for subjects
    denied in Yes/No):

      1. Multi expert consensus  (n>=1, agreement>=67%)   -> multi_expert
      2. Multi crowd consensus   (n>=3, agreement>=67%)   -> multi_consensus
         (either multi step above overrides a yes/no denial;
          a "not sure" consensus label -> voted_review instead)
      3. Yes/No expert confirm   (n>=1, >=75% yes)         -> confirm_expert
      4. Yes/No crowd confirm    (n>=5, >=75% yes)         -> confirm_pred
         (only if no multi consensus, and expert didn't deny)
      5. Denied in yes/no (expert or crowd, >=75% no), AND
         it has since been retired from every Multi-choice
         workflow it was sent to without reaching Multi
         consensus there either (dead end)                 -> stalled_no_consensus
      6. Denied in yes/no, but Multi-choice hasn't reached a
         verdict yet (still active there, or not yet queued
         into a Multi subject set) — this is expected
         pipeline routing, not a problem                    -> pending_multi_review
      7. Never confirmed/denied in Yes/No, and retired from
         every Yes/No workflow it was sent to (hit the
         classification-count cap before reaching either
         threshold)                                         -> stalled_no_consensus
      8. Otherwise (still active on Zooniverse)              -> needs_more_votes

    consensus_status values 1-4 correspond to Toolbox Verified=TRUE. Of the
    rest, only needs_more_votes/pending_multi_review are subjects that could
    still resolve on their own with more Zooniverse votes — voted_review and
    stalled_no_consensus are dead ends that need a human decision in Toolbox.
    Being denied in Yes/No is deliberately NOT treated as "needs review" by
    itself: a "No" consensus is exactly what's supposed to route a subject
    into the Multi-choice workflow next, so it only becomes a review case
    once that next stage has also stalled without a verdict.

    Returns one row per subject_ids seen in `flat`, so callers can merge this
    back onto any subject-level view (subject summary, source image rollups,
    transect completion).
    """
    def _yn_votes(ids: set) -> pd.DataFrame:
        g = flat[flat["workflow_id"].isin(ids)]
        cols = ["subject_ids", "n", "yes_frac", "no_frac"]
        if g.empty:
            return pd.DataFrame(columns=cols).set_index("subject_ids")
        ans = g["answer"].astype(str).str.strip().str.lower()
        is_yes = ans.eq("yes")
        is_no  = ans.str.contains("no") & ~is_yes
        tmp = pd.DataFrame({"subject_ids": g["subject_ids"].values,
                            "is_yes": is_yes.values, "is_no": is_no.values})
        agg = tmp.groupby("subject_ids").agg(n=("is_yes", "size"),
                                             yes=("is_yes", "sum"),
                                             no=("is_no", "sum"))
        agg["yes_frac"] = agg["yes"] / agg["n"]
        agg["no_frac"]  = agg["no"] / agg["n"]
        return agg[["n", "yes_frac", "no_frac"]]

    def _multi_votes(ids: set) -> pd.DataFrame:
        g = flat[flat["workflow_id"].isin(ids)]
        g = g[g["answer"].astype(str).str.len() > 0]
        cols = ["subject_ids", "n", "top_label", "top_count", "agreement"]
        if g.empty:
            return pd.DataFrame(columns=cols).set_index("subject_ids")
        counts = (g.groupby(["subject_ids", "answer"]).size()
                   .reset_index(name="count"))
        counts = counts.sort_values(["subject_ids", "count", "answer"],
                                    ascending=[True, False, True])
        top = (counts.drop_duplicates("subject_ids", keep="first")
                     .set_index("subject_ids")
                     .rename(columns={"answer": "top_label", "count": "top_count"}))
        n = g.groupby("subject_ids").size().rename("n")
        out = top.join(n)
        out["agreement"] = out["top_count"] / out["n"]
        return out[["n", "top_label", "top_count", "agreement"]]

    yn     = _yn_votes(YESNO_WORKFLOW_IDS)
    yn_exp = _yn_votes(YESNO_EXPERT_WORKFLOW_IDS)
    m      = _multi_votes(MULTI_WORKFLOW_IDS)
    m_exp  = _multi_votes(MULTI_EXPERT_WORKFLOW_IDS)

    ids = pd.Index(flat["subject_ids"].dropna().unique(), name="subject_ids")
    out = pd.DataFrame(index=ids)

    def _num(series, default=0.0):
        return pd.to_numeric(series, errors="coerce").fillna(default)

    out["yn_n"]            = _num(yn.reindex(ids)["n"]).astype(int)
    out["yn_yes_frac"]     = _num(yn.reindex(ids)["yes_frac"])
    out["yn_no_frac"]      = _num(yn.reindex(ids)["no_frac"])
    out["yn_exp_n"]        = _num(yn_exp.reindex(ids)["n"]).astype(int)
    out["yn_exp_yes_frac"] = _num(yn_exp.reindex(ids)["yes_frac"])
    out["yn_exp_no_frac"]  = _num(yn_exp.reindex(ids)["no_frac"])
    out["m_n"]             = _num(m.reindex(ids)["n"]).astype(int)
    out["m_top_label"]     = m.reindex(ids)["top_label"]
    out["m_agreement"]     = _num(m.reindex(ids)["agreement"])
    out["m_exp_n"]         = _num(m_exp.reindex(ids)["n"]).astype(int)
    out["m_exp_top_label"] = m_exp.reindex(ids)["top_label"]
    out["m_exp_agreement"] = _num(m_exp.reindex(ids)["agreement"])

    yn_confirm     = (out["yn_n"]     >= YESNO_AGREE_MIN_N) & (out["yn_yes_frac"]     >= YESNO_AGREE_MIN_FRAC)
    yn_deny        = (out["yn_n"]     >= YESNO_AGREE_MIN_N) & (out["yn_no_frac"]      >= YESNO_AGREE_MIN_FRAC)
    yn_exp_confirm = (out["yn_exp_n"] >= EXPERT_MIN_N)      & (out["yn_exp_yes_frac"] >= YESNO_AGREE_MIN_FRAC)
    yn_exp_deny    = (out["yn_exp_n"] >= EXPERT_MIN_N)      & (out["yn_exp_no_frac"]  >= YESNO_AGREE_MIN_FRAC)

    m_consensus     = (out["m_n"]     >= MULTI_AGREE_MIN_N) & (out["m_agreement"]     >= MULTI_AGREE_MIN_FRAC)
    m_exp_consensus = (out["m_exp_n"] >= EXPERT_MIN_N)      & (out["m_exp_agreement"] >= MULTI_AGREE_MIN_FRAC)

    m_not_sure     = out["m_top_label"].astype(str).str.contains(NOT_SURE_RE)
    m_exp_not_sure = out["m_exp_top_label"].astype(str).str.contains(NOT_SURE_RE)

    yn_denied = yn_deny | yn_exp_deny

    # ── Retirement, scoped per pipeline stage ───────────────────────────
    # "Retired everywhere" needs to be checked separately for the Yes/No
    # stage and the Multi-choice stage, because a subject denied in Yes/No
    # is *expected* to still be active (or not yet queued) in Multi-choice —
    # that's the pipeline working normally, not a dead end. A stage counts
    # as exhausted only when the subject appeared in it and is retired in
    # every workflow of that stage it appeared in.
    def _stage_retirement(workflow_ids: set):
        stage = flat[flat["workflow_id"].isin(workflow_ids)]
        if stage.empty:
            empty = pd.Series(False, index=ids)
            return empty, empty.copy()
        retired_by_wf = stage.groupby(["subject_ids", "workflow_id"])["is_retired"].first()
        has_data       = retired_by_wf.groupby("subject_ids").size().reindex(ids, fill_value=0).gt(0)
        retired_all    = retired_by_wf.groupby("subject_ids").all().reindex(ids, fill_value=False)
        return has_data, has_data & retired_all

    yn_ids_set    = YESNO_WORKFLOW_IDS | YESNO_EXPERT_WORKFLOW_IDS
    multi_ids_set = MULTI_WORKFLOW_IDS | MULTI_EXPERT_WORKFLOW_IDS

    yn_has_data, yn_stage_exhausted       = _stage_retirement(yn_ids_set)
    multi_has_data, multi_stage_exhausted = _stage_retirement(multi_ids_set)
    overall_has_data, overall_exhausted   = _stage_retirement(yn_ids_set | multi_ids_set)

    out["zooniverse_active"]             = overall_has_data & ~overall_exhausted
    out["zooniverse_retired_everywhere"] = overall_has_data & overall_exhausted

    denied_multi_stalled = yn_denied & multi_has_data & multi_stage_exhausted
    denied_pending       = yn_denied & ~denied_multi_stalled
    yn_only_stalled       = (~yn_denied) & yn_has_data & yn_stage_exhausted

    out["consensus_status"] = np.select(
        [m_exp_consensus & m_exp_not_sure,
         m_exp_consensus,
         m_consensus & m_not_sure,
         m_consensus,
         yn_exp_confirm,
         yn_confirm & ~yn_exp_deny,
         denied_multi_stalled,
         denied_pending,
         yn_only_stalled],
        ["voted_review",
         "multi_expert",
         "voted_review",
         "multi_consensus",
         "confirm_expert",
         "confirm_pred",
         "stalled_no_consensus",
         "pending_multi_review",
         "stalled_no_consensus"],
        default="needs_more_votes"
    )

    out["zooniverse_verified"] = out["consensus_status"].isin(
        ["multi_expert", "multi_consensus", "confirm_expert", "confirm_pred"])
    outcome = out["consensus_status"].map(CONSENSUS_OUTCOME_MAP)
    out["consensus_outcome"] = outcome.map(lambda t: t[0] if isinstance(t, tuple) else "Zooniverse — Needs More Votes")
    out["consensus_detail"]  = outcome.map(lambda t: t[1] if isinstance(t, tuple) else "Insufficient votes")

    # Refine detail text with the specific reason for the two multi-stage buckets.
    stalled_after_denial = out["consensus_status"].eq("stalled_no_consensus") & denied_multi_stalled
    out.loc[stalled_after_denial, "consensus_detail"] = (
        "Denied in Yes/No, then retired from the Multi-choice workflow without "
        "reaching consensus there either; will not receive more votes automatically")

    pending = out["consensus_status"].eq("pending_multi_review")
    multi_votes_so_far = out["m_n"] + out["m_exp_n"]
    has_votes = pending & (multi_votes_so_far > 0)
    no_votes  = pending & ~has_votes
    out.loc[has_votes, "consensus_detail"] = (
        "Denied in Yes/No; awaiting Multi-choice consensus on Zooniverse ("
        + multi_votes_so_far[has_votes].astype(int).astype(str) + " vote(s) so far)")
    out.loc[no_votes, "consensus_detail"] = (
        "Denied in Yes/No; not yet classified in the Multi-choice workflow")

    return out.reset_index()


# ============================================================
# LOAD & FLATTEN
# ============================================================
def load_and_flatten(csv_path: str,
                     workflow_id: int | None,
                     source_image_filter: str | None,
                     transect_id_filter: str | None = None,
                     date_from: str | None = None,
                     date_to: str | None = None) -> pd.DataFrame:
    """
    Load the raw export CSV and join all parsed columns into a single flat table.
    Applies --workflow-id and --source-image filters if provided.
    """
    log.info(f"Loading {csv_path}…")
    raw = pd.read_csv(csv_path, low_memory=False)
    log.info(f"Loaded {len(raw):,} rows, {len(raw.columns)} columns")

    # Workflow filter
    if workflow_id is not None:
        raw = raw[raw["workflow_id"] == workflow_id].copy()
        log.info(f"After workflow filter ({workflow_id}): {len(raw):,} rows")
        if raw.empty:
            log.error(f"No rows found for workflow_id={workflow_id}. "
                      f"Available: {list(raw['workflow_id'].unique())}")
            sys.exit(1)

    log.info("Parsing JSON columns…")
    subj_df = parse_subject_data(raw["subject_data"])
    meta_df = parse_metadata(raw["metadata"])
    answers  = parse_annotations(raw["annotations"], raw["workflow_id"])

    flat = pd.concat([
        raw[["classification_id", "user_name", "user_id",
             "workflow_id", "workflow_name", "workflow_version",
             "created_at", "subject_ids"]].reset_index(drop=True),
        subj_df.reset_index(drop=True),
        meta_df.reset_index(drop=True),
        answers.rename("answer").reset_index(drop=True),
    ], axis=1)

    # Derived columns
    user_name_norm = flat["user_name"].fillna("").astype(str).str.strip().str.lower()
    flat["is_anonymous"]    = user_name_norm.str.startswith("not-logged-in")
    flat["created_at"]      = pd.to_datetime(flat["created_at"], utc=True, errors="coerce")

    # Source-image filter (applied after parsing)
    if source_image_filter is not None:
        flat = flat[flat["source_image"] == source_image_filter].copy()
        log.info(f"After source_image filter ('{source_image_filter}'): {len(flat):,} rows")
        if flat.empty:
            log.error(f"No rows matched source_image='{source_image_filter}'.")
            sys.exit(1)

    # Transect ID filter — accepts comma-separated list
    if transect_id_filter is not None:
        transect_ids = [t.strip() for t in transect_id_filter.split(",") if t.strip()]
        flat = flat[flat["transect_id"].isin(transect_ids)].copy()
        log.info(f"After transect_id filter ({transect_ids}): {len(flat):,} rows")
        if flat.empty:
            log.error(f"No rows matched transect_id in {transect_ids}.")
            sys.exit(1)

    # Date range filter on classification created_at
    if date_from is not None:
        try:
            dt_from = pd.Timestamp(date_from, tz="UTC")
            flat = flat[flat["created_at"] >= dt_from].copy()
            log.info(f"After date_from filter (>= {date_from}): {len(flat):,} rows")
        except Exception:
            log.error(f"Invalid date_from value: '{date_from}'. Use YYYY-MM-DD.")
            sys.exit(1)

    if date_to is not None:
        try:
            dt_to = pd.Timestamp(date_to, tz="UTC") + pd.Timedelta(days=1)
            flat = flat[flat["created_at"] < dt_to].copy()
            log.info(f"After date_to filter (<= {date_to}): {len(flat):,} rows")
        except Exception:
            log.error(f"Invalid date_to value: '{date_to}'. Use YYYY-MM-DD.")
            sys.exit(1)

    if (date_from or date_to) and flat.empty:
        log.error("No rows remain after date filter.")
        sys.exit(1)

    log.info(f"Flat table: {len(flat):,} rows, {len(flat.columns)} columns")

    if workflow_id is not None or date_from is not None or date_to is not None:
        log.warning("Workflow/date filters are active — the Verified / Needs Toolbox "
                    "Review status below is computed only from the votes present in "
                    "this filtered export and may not reflect the true outcome once "
                    "all workflows and classifications are counted.")

    log.info("Computing Zooniverse consensus / Toolbox review status per subject "
             "(same thresholds as zooni_to_toolbox_annot.py)…")
    consensus = compute_consensus(flat)
    flat = flat.merge(
        consensus[["subject_ids", "consensus_status", "zooniverse_verified",
                   "consensus_outcome", "consensus_detail",
                   "zooniverse_active", "zooniverse_retired_everywhere"]],
        on="subject_ids", how="left")

    return flat


# ============================================================
# SUMMARY BUILDERS
# ============================================================
def build_overview(flat: pd.DataFrame, filters: dict) -> pd.DataFrame:
    wf_list = (flat[["workflow_id", "workflow_name"]]
               .drop_duplicates()
               .assign(combined=lambda d: d["workflow_id"].astype(str)
                       + " — " + d["workflow_name"])
               ["combined"].tolist())

    subjects = flat.drop_duplicates("subject_ids")
    outcome_counts = subjects["consensus_outcome"].value_counts()

    rows = [
        ("Export file rows analysed",   len(flat)),
        ("Unique subjects",             flat["subject_ids"].nunique()),
        ("Unique classifiers (named)",  flat.loc[~flat["is_anonymous"], "user_name"].nunique()),
        ("Non-logged-in classifications", flat["is_anonymous"].sum()),
        ("Non-logged-in %",             f"{flat['is_anonymous'].mean()*100:.1f}%"),
        ("Retired subjects (all subject sets exhausted)", int(subjects["zooniverse_retired_everywhere"].sum())),
        ("Active subjects (still on Zooniverse somewhere)", int(subjects["zooniverse_active"].sum())),
        ("Verified — Yes/No (Zooniverse consensus)",     int(subjects["consensus_status"].isin(["confirm_pred", "confirm_expert"]).sum())),
        ("Verified — Multi-choice (Zooniverse consensus)", int(subjects["consensus_status"].isin(["multi_consensus", "multi_expert"]).sum())),
        ("Zooniverse — Needs More Votes (still active, may resolve on its own)",
         int(outcome_counts.get("Zooniverse — Needs More Votes", 0))),
        ("—   of which denied in Yes/No, awaiting Multi-choice",
         int((subjects["consensus_status"] == "pending_multi_review").sum())),
        ("Needs Toolbox Review (Zooniverse exhausted, no verdict — bring to Toolbox)",
         int(outcome_counts.get("Needs Toolbox Review", 0))),
        ("—   of which voted 'not sure' in Multi-choice",
         int((subjects["consensus_status"] == "voted_review").sum())),
        ("—   of which stalled (retired, no consensus reached)",
         int((subjects["consensus_status"] == "stalled_no_consensus").sum())),
        ("Confirmation thresholds (match zooni_to_toolbox_annot.py)",
         f"Yes/No: n≥{YESNO_AGREE_MIN_N} & ≥{YESNO_AGREE_MIN_FRAC:.0%} yes | "
         f"Yes/No expert: n≥{EXPERT_MIN_N} & ≥{YESNO_AGREE_MIN_FRAC:.0%} yes | "
         f"Multi: n≥{MULTI_AGREE_MIN_N} & ≥{MULTI_AGREE_MIN_FRAC:.0%} agreement | "
         f"Multi expert: n≥{EXPERT_MIN_N} & ≥{MULTI_AGREE_MIN_FRAC:.0%} agreement"),
        ("Date range (first)",          str(flat["created_at"].min())[:19] if flat["created_at"].notna().any() else "—"),
        ("Date range (last)",           str(flat["created_at"].max())[:19] if flat["created_at"].notna().any() else "—"),
        ("Workflows in this file",      ", ".join(wf_list)),
        ("— filter: workflow_id",       str(filters.get("workflow_id") or "all")),
        ("— filter: source_image",      str(filters.get("source_image") or "all")),
        ("— filter: transect_id",       str(filters.get("transect_id") or "all")),
        ("Report generated",            datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def build_workflow_summary(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for wf_id, g in flat.groupby("workflow_id"):
        subjects = g.drop_duplicates("subject_ids")
        retired  = subjects[subjects["is_retired"]]
        rows.append({
            "Workflow ID":                  wf_id,
            "Workflow Name":                g["workflow_name"].iloc[0],
            "Workflow Version(s)":          ", ".join(g["workflow_version"].astype(str).unique()),
            "Total Classifications":        len(g),
            "Unique Subjects":              g["subject_ids"].nunique(),
            "Retired Subjects":             len(retired),
            "Not Retired":                  len(subjects) - len(retired),
            "Retired — consensus":          (retired["retirement_reason"] == "consensus").sum(),
            "Retired — classification_count": (retired["retirement_reason"] == "classification_count").sum(),
            "Retired — other":              (~retired["retirement_reason"].isin(
                                                ["consensus", "classification_count"])).sum()
                                            if len(retired) else 0,
            "Avg classifications/subject":  f"{g.groupby('subject_ids').size().mean():.2f}",
            "Non-logged-in classifications": g["is_anonymous"].sum(),
            "Non-logged-in %":            f"{g['is_anonymous'].mean()*100:.1f}%",
            "Unique named users":           g.loc[~g["is_anonymous"], "user_name"].nunique(),
            "Desktop classifications":      (g["device"] == "desktop").sum(),
            "Mobile classifications":       (g["device"] == "mobile").sum(),
            "Users in a group":             g["in_user_group"].sum(),
        })
    return pd.DataFrame(rows)


def build_subject_summary(flat: pd.DataFrame) -> pd.DataFrame:
    """
    One row per unique subject (image patch) with retirement info, vote
    counts, and the classification outcome (see compute_consensus).

    "Outcome" is one of three things — and only one of them is a Toolbox
    concern:
      - Verified …             : Zooniverse crowd/expert consensus decided
                                  it. Nothing to do; this is what will be
                                  written into Toolbox's Verified=TRUE field.
      - Zooniverse — Needs More Votes : still active, may resolve on its own.
      - Needs Toolbox Review    : Zooniverse produced no verdict and won't
                                  produce one automatically — bring this
                                  subject into Toolbox for a human decision.

    "Zooniverse Status" is "Retired" only if the subject has been retired in
    *every* tracked workflow/subject set it was sent to (won't receive any
    more votes); otherwise "Active" (still circulating somewhere and could
    still gain votes). A subject can show "Zooniverse — Needs More Votes" +
    "Active" (genuinely still pending) or "Needs Toolbox Review" + "Retired"
    (Zooniverse is done with it, but our threshold was never reached) —
    those are the two ends of the same axis.
    """
    rows = []
    for subj_id, g in flat.groupby("subject_ids"):
        r = g.iloc[0]
        # "any row retired" is just informational context for the reason/date
        # below — the authoritative retirement signal is zooniverse_retired_everywhere.
        retired_any = bool(g["is_retired"].any())
        retired_everywhere = bool(r["zooniverse_retired_everywhere"])

        top_answer = g["answer"].mode().iloc[0] if not g["answer"].empty else ""

        rows.append({
            "Subject ID":                       subj_id,
            "Filename":                         r["filename"],
            "Source Image":                     r["source_image"],
            "Transect ID":                      r["transect_id"],
            "Model Prediction Code":            r["model_pred_code"],
            "Model Prediction Name":            r["model_pred_name"],
            "Total Classifications":            len(g),
            "Top Answer (most votes)":          top_answer,
            "Outcome":                          r["consensus_outcome"],
            "Outcome Detail":                   r["consensus_detail"],
            "Zooniverse Status":                "Retired (all subject sets)" if retired_everywhere else "Active",
            "Retirement Reason":                r["retirement_reason"] if retired_any else "—",
            "Classifications at Retirement":    r["classifications_count_at_retire"] if retired_any else "—",
            "Retired At":                       str(r["retired_at"])[:19] if retired_any and r["retired_at"] else "—",
            "Workflow(s)":                      ", ".join(g["workflow_id"].astype(str).unique()),
        })
    df = pd.DataFrame(rows)
    # Sort: Needs Toolbox Review first (this is the only bucket needing human
    # action), then still-pending-on-Zooniverse, then Verified last.
    outcome_priority = {"Needs Toolbox Review": 0, "Zooniverse — Needs More Votes": 1}
    df["_sort"] = df["Outcome"].map(lambda o: outcome_priority.get(o, 2))
    df = df.sort_values(["_sort", "Total Classifications"],
                        ascending=[True, False]).drop(columns=["_sort"]).reset_index(drop=True)
    return df


def build_user_summary(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for user, g in flat.groupby("user_name"):
        rows.append({
            "User Name":                user,
            "Is Non-logged-in":         g["is_anonymous"].iloc[0],
            "Total Classifications":    len(g),
            "Workflows Contributed To": ", ".join(g["workflow_id"].astype(str).unique()),
            "Unique Subjects Seen":     g["subject_ids"].nunique(),
            "In a User Group":          g["in_user_group"].any(),
            "Desktop":                  (g["device"] == "desktop").sum(),
            "Mobile":                   (g["device"] == "mobile").sum(),
            "Avg Duration (s)":         f"{g['duration_s'].mean():.1f}" if g["duration_s"].notna().any() else "—",
            "Min Duration (s)":         f"{g['duration_s'].min():.1f}" if g["duration_s"].notna().any() else "—",
            "First Classification":     str(g["created_at"].min())[:19],
            "Last Classification":      str(g["created_at"].max())[:19],
        })
    df = pd.DataFrame(rows).sort_values("Total Classifications", ascending=False).reset_index(drop=True)
    return df


def build_answer_breakdown(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for wf_id, g in flat.groupby("workflow_id"):
        wf_name   = g["workflow_name"].iloc[0]
        total     = len(g)
        vc        = g["answer"].value_counts()
        for answer, count in vc.items():
            rows.append({
                "Workflow ID":          wf_id,
                "Workflow Name":        wf_name,
                "Answer":               answer,
                "Count":                count,
                "% of Workflow Total":  f"{count/total*100:.1f}%",
            })
    return pd.DataFrame(rows)


def build_source_image_summary(flat: pd.DataFrame) -> pd.DataFrame:
    """
    Per source (transect) image: how many patches (subjects) are retired on
    the Zooniverse platform, and — separately — how many of those patches
    have actually been confirmed via the Yes/No or Multi-choice workflow
    consensus rules (Verified), vs are still active and waiting on more
    volunteer votes (Zooniverse's job, not Toolbox's), vs have exhausted
    Zooniverse without a verdict and need a human to review them in Toolbox.
    Platform retirement and Verified status can disagree (e.g. a subject can
    be platform-retired on classification_count before reaching our
    agreement threshold — that's exactly the "Needs Toolbox Review" case).
    """
    rows = []
    for src, g in flat.groupby("source_image"):
        subjects  = g.drop_duplicates("subject_ids")
        status    = subjects["consensus_status"]
        rows.append({
            "Source Image":                 src,
            "Total Subjects (patches)":     len(subjects),
            "Retired (all subject sets)":   int(subjects["zooniverse_retired_everywhere"].sum()),
            "Active on Zooniverse":         int(subjects["zooniverse_active"].sum()),
            "Verified — Yes/No":            int(status.isin(["confirm_pred", "confirm_expert"]).sum()),
            "Verified — Multi-choice":      int(status.isin(["multi_consensus", "multi_expert"]).sum()),
            "Zooniverse — Needs More Votes": int(status.isin(["needs_more_votes", "pending_multi_review"]).sum()),
            "— Denied in Yes/No, awaiting Multi-choice": int((status == "pending_multi_review").sum()),
            "Needs Toolbox Review":         int(status.isin(["voted_review", "stalled_no_consensus"]).sum()),
            "— Voted Not Sure (Multi)":     int((status == "voted_review").sum()),
            "— Stalled (retired, no consensus)": int((status == "stalled_no_consensus").sum()),
            "Total Classifications":        len(g),
            "Avg Classifications/Subject":  f"{len(g)/len(subjects):.2f}" if len(subjects) else "—",
            "Unique Users":                 g.loc[~g["is_anonymous"], "user_name"].nunique(),
            "Non-logged-in Classifications": g["is_anonymous"].sum(),
            "Workflow(s)":                  ", ".join(g["workflow_id"].astype(str).unique()),
        })
    df = pd.DataFrame(rows).sort_values("Total Subjects (patches)", ascending=False).reset_index(drop=True)
    return df


def build_time_stats(flat: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for wf_id, g in flat.groupby("workflow_id"):
        d = g["duration_s"].dropna()
        if d.empty:
            continue
        rows.append({
            "Workflow ID":          wf_id,
            "Workflow Name":        g["workflow_name"].iloc[0],
            "N (with timing)":      len(d),
            "Mean (s)":             round(d.mean(), 1),
            "Median (s)":           round(d.median(), 1),
            "Std Dev (s)":          round(d.std(), 1),
            "Min (s)":              round(d.min(), 1),
            "Max (s)":              round(d.max(), 1),
            "< 3 s (likely rushed)": (d < 3).sum(),
            "< 3 s %":              f"{(d < 3).mean()*100:.1f}%",
            "> 60 s":               (d > 60).sum(),
        })
    return pd.DataFrame(rows)


def build_transect_completion(flat: pd.DataFrame) -> pd.DataFrame:
    """
    One row per transect showing pipeline completion, using the same vote
    thresholds as zooni_to_toolbox_annot.py (see compute_consensus) rather
    than Zooniverse's own platform retirement flag.

    Groups by transect_id when present in subject metadata (stamped by
    import_subjects.py / patch_subject_metadata.py). Falls back to
    source_image for older subjects that pre-date that field.

    A subject is *Resolved* once it has a determination — either Verified
    (Zooniverse crowd/expert consensus, Yes/No or Multi-choice) or flagged
    Needs Toolbox Review (crowd voted "not sure" in Multi-choice, or
    Stalled: retired from every workflow of a pipeline stage without ever
    reaching our threshold there — Zooniverse is done and produced no
    verdict, so a human needs to open Toolbox and decide). Being denied in
    Yes/No is NOT by itself a review case — it's the expected trigger for
    the Multi-choice stage, so those subjects count as *Pending* (see
    "awaiting Multi-choice") right along with subjects still collecting
    Yes/No votes — Zooniverse's job, not Toolbox's — unless Multi-choice has
    also stalled without a verdict.

    "Retired (all subject sets)" is reported alongside for reference, since
    the platform can retire a subject (e.g. on classification_count) before
    or after our agreement threshold is reached — the two do not always
    agree, which is exactly what the Stalled bucket above is for.
    """
    flat = flat.copy()
    flat["_transect_key"] = flat["transect_id"].where(
        flat["transect_id"].notna() & (flat["transect_id"].astype(str).str.strip() != ""),
        other=flat["source_image"]
    )

    yn_expert_subjects    = set(flat.loc[flat["workflow_id"].isin(YESNO_EXPERT_WORKFLOW_IDS), "subject_ids"])
    multi_expert_subjects = set(flat.loc[flat["workflow_id"].isin(MULTI_EXPERT_WORKFLOW_IDS), "subject_ids"])

    subj = flat.drop_duplicates("subject_ids")[
        ["subject_ids", "_transect_key", "is_retired", "consensus_status",
         "zooniverse_retired_everywhere"]
    ].copy()
    subj = subj[subj["_transect_key"].notna() & (subj["_transect_key"].astype(str).str.strip() != "")]
    subj["sent_to_yn_expert"]    = subj["subject_ids"].isin(yn_expert_subjects)
    subj["sent_to_multi_expert"] = subj["subject_ids"].isin(multi_expert_subjects)

    rows = []
    for key, g in subj.groupby("_transect_key"):
        status_col   = g["consensus_status"]
        total        = len(g)
        confirmed_yn = int(status_col.isin(["confirm_pred", "confirm_expert"]).sum())
        confirmed_m  = int(status_col.isin(["multi_consensus", "multi_expert"]).sum())
        not_sure     = int((status_col == "voted_review").sum())
        stalled      = int((status_col == "stalled_no_consensus").sum())
        needs_review = not_sure + stalled
        awaiting_multi = int((status_col == "pending_multi_review").sum())
        needs_votes  = int((status_col == "needs_more_votes").sum())
        pending      = needs_votes + awaiting_multi
        resolved     = confirmed_yn + confirmed_m + needs_review
        pct          = resolved / total * 100 if total > 0 else 0.0
        status       = "Complete" if resolved == total else "In Progress"

        rows.append({
            "Transect":                        key,
            "Status":                          status,
            "Resolved %":                      round(pct, 1),
            "Total Subjects":                  total,
            "Resolved":                        resolved,
            "Zooniverse — Needs More Votes":   pending,
            "— Awaiting Yes/No votes":         needs_votes,
            "— Denied in Yes/No, awaiting Multi-choice": awaiting_multi,
            # ── Outcome breakdown ──────────────────────────────
            "Confirmed — Yes/No":              confirmed_yn,
            "Confirmed — Multi-choice":        confirmed_m,
            "Needs Toolbox Review":            needs_review,
            "— Voted Not Sure (Multi)":        not_sure,
            "— Stalled (retired, no consensus)": stalled,
            # ── Expert review ─────────────────────────────────
            "Sent to Expert Yes/No":           int(g["sent_to_yn_expert"].sum()),
            "Sent to Expert Multi-choice":     int(g["sent_to_multi_expert"].sum()),
            # ── Zooniverse platform retirement (reference only) ──
            "Retired (all subject sets)":      int(g["zooniverse_retired_everywhere"].sum()),
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    # Sort: In Progress first, then Complete; within each group by resolved % desc
    status_order = {"In Progress": 0, "Complete": 1}
    df["_sort"] = df["Status"].map(status_order)
    df = (df.sort_values(["_sort", "Resolved %"], ascending=[True, False])
            .drop(columns=["_sort"])
            .reset_index(drop=True))
    return df


# ============================================================
# EXCEL FORMATTING
# ============================================================
def _thin():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    thin = Side(style="thin",   color="BDBDBD")
    med  = Side(style="medium", color=TEAL_MID)
    return Border(left=thin, right=thin, top=thin, bottom=med)

def _fill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def _font(bold=False, size=10, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def style_sheet(ws, df: pd.DataFrame,
                header_color: str = TEAL_MID,
                title: str = "",
                col_widths: dict | None = None):
    """Write df to ws with consistent header + alternating row styling."""
    ws.sheet_view.showGridLines = False

    start_row = 1
    if title:
        ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
        c = ws["A1"]
        c.value     = title
        c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
        c.fill      = _fill(TEAL_DARK)
        c.alignment = _center()
        ws.row_dimensions[1].height = 26
        start_row = 2

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=col_idx, value=col_name)
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill(header_color)
        c.alignment = _center(wrap=True)
        c.border    = _thick_bottom()
    ws.row_dimensions[start_row].height = 28

    # Data rows
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        bg = TEAL_PALE if r_idx % 2 == 0 else WHITE
        for c_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=r_idx, column=c_idx,
                        value=None if (isinstance(val, float) and pd.isna(val)) else val)
            c.font      = _font()
            c.fill      = _fill(bg)
            c.alignment = _left(wrap=False)
            c.border    = _thin()
        ws.row_dimensions[r_idx].height = 16

    # Column widths
    default_widths = {i: max(len(str(col)) + 4, 12)
                      for i, col in enumerate(df.columns, start=1)}
    # Widen based on data sample
    for r_idx, row_data in enumerate(df.head(40).itertuples(index=False)):
        for c_idx, val in enumerate(row_data, start=1):
            w = min(len(str(val)) + 3, 50)
            if w > default_widths.get(c_idx, 0):
                default_widths[c_idx] = w

    for c_idx, width in default_widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    if col_widths:
        for col_name, width in col_widths.items():
            if col_name in df.columns:
                c_idx = list(df.columns).index(col_name) + 1
                ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


# ============================================================
# WRITE REPORT
# ============================================================
def write_report(sheets: dict[str, pd.DataFrame], output_path: Path,
                 header_colors: dict[str, str] | None = None):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    colors = header_colors or {}
    titles = {
        "Overview":             "📊  Classification Export — Overview",
        "Transect Completion":  "🌊  Transect Completion — Pipeline Progress",
        "Workflow Summary":     "📋  Workflow Summary",
        "Subject Summary":      "🔎  Subject Summary",
        "User Summary":         "👤  User Summary",
        "Answer Breakdown":     "✅  Answer Breakdown",
        "Source Image":         "🗂️  Source Image Summary",
        "Time Stats":           "⏱️  Classification Time Statistics",
    }

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name)
        style_sheet(
            ws, df,
            header_color=colors.get(sheet_name, TEAL_MID),
            title=titles.get(sheet_name, sheet_name),
        )
        log.info(f"  Sheet '{sheet_name}': {len(df)} rows")

    wb.save(output_path)
    log.info(f"Report saved → {output_path}")


# ============================================================
# ENTRY POINT
# ============================================================
def main():
    args = get_args_via_gui()

    csv_path = Path(args.export_csv)
    if not csv_path.is_file():
        log.error(f"Export CSV not found: {csv_path}")
        sys.exit(1)

    filters = {
        "workflow_id":  args.workflow_id,
        "source_image": args.source_image,
        "transect_id":  args.transect_id,
        "date_from":    args.date_from,
        "date_to":      args.date_to,
    }

    flat = load_and_flatten(
        str(csv_path),
        workflow_id=args.workflow_id,
        source_image_filter=args.source_image,
        transect_id_filter=args.transect_id,
        date_from=args.date_from,
        date_to=args.date_to,
    )

    log.info("Building summary sheets…")
    sheets = {
        "Overview":             build_overview(flat, filters),
        "Transect Completion":  build_transect_completion(flat),
        "Workflow Summary":     build_workflow_summary(flat),
        "Subject Summary":      build_subject_summary(flat),
        "User Summary":         build_user_summary(flat),
        "Answer Breakdown":     build_answer_breakdown(flat),
        "Source Image":         build_source_image_summary(flat),
        "Time Stats":           build_time_stats(flat),
    }

    sheet_colors = {
        "Overview":             TEAL_DARK,
        "Transect Completion":  GREEN_DARK,
        "Workflow Summary":     TEAL_MID,
        "Subject Summary":      TEAL_MID,
        "User Summary":         "5D6D7E",
        "Answer Breakdown":     GREEN_DARK,
        "Source Image":         ORANGE,
        "Time Stats":           "7D3C98",
    }

    # Build output filename
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str   = datetime.now().strftime("%Y%m%d_%H%M")
    wf_suffix  = f"_wf{args.workflow_id}"      if args.workflow_id  else ""
    si_suffix  = f"_{args.source_image}"       if args.source_image else ""
    ti_suffix  = f"_{args.transect_id}"        if args.transect_id  else ""
    out_path   = output_dir / f"classification_report{wf_suffix}{si_suffix}{ti_suffix}_{date_str}.xlsx"

    write_report(sheets, out_path, header_colors=sheet_colors)

    if args.open_report:
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(out_path))
            elif sys.platform == "darwin":
                subprocess.run(["open", str(out_path)], check=False)
            else:
                subprocess.run(["xdg-open", str(out_path)], check=False)
        except Exception as e:
            log.warning(f"Could not open report automatically: {e}")

    # Console summary
    log.info("=" * 60)
    log.info("✅  Report complete!")
    log.info(f"    Rows analysed     : {len(flat):,}")
    log.info(f"    Unique subjects   : {flat['subject_ids'].nunique():,}")
    log.info(f"    Retired subjects  : {flat.drop_duplicates('subject_ids')['is_retired'].sum():,}")
    log.info(f"    Unique workflows  : {flat['workflow_id'].nunique()}")
    log.info(f"    Output            : {out_path}")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
